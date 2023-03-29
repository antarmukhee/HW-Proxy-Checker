__author__      = "Nandan Sharma"

import tomllib
import re
from openpyxl import load_workbook, Workbook
from pprint import pprint
import deepdiff
from openpyxl import Workbook as openpyxl_wb

# global variables
config_file = r'config.toml'
config_dict = tomllib.load(open(config_file, "rb"))

config_dict_str = str(config_dict)
config_cabinet = config_dict['cabinet']
config_station = config_dict['station']
config_slot = config_dict['slot']
config_channel = config_dict['channel']
config_rt = config_dict['rt']

file_hw_assign = config_dict['file_hw_assign']
file_t3000_1 = config_dict['file_t3000_1']
file_t3000_2 = config_dict['file_t3000_2']
if file_t3000_2.strip() == '': file_t3000_2 = None
file_diff = config_dict['file_diff']
prefix = str(config_dict['prefix'])

file_hw_assign_out = None # r'.\result\hw_assign_out.xlsx'
file_t3000_assign_out = None # r'.\result\t3000_assign_out.xlsx'

tagname_t3k_hw_proxy_format = config_dict['tagname_t3k_hw_proxy']['format']
def tagname_t3k_hw_proxy(cabinet, station, slot):
    return eval(tagname_t3k_hw_proxy_format)

def is_equal_float(x1, x2):
    try:
        return float(x1) == float(x2)
    except ValueError:
        return False

def io_data_hw_assignment():
    hw_assign_fields = re.findall("\@(.*?)\@", config_dict_str)
    hw_assign_fields = list(set(hw_assign_fields) - set(["CABINET", "STATION", "SLOT", "CHANNEL", "RT", "TYPE", "MLFB", "change_token", "KKS", "SIGDEF"]))
    print("HW_assignment_list fields which will be looked up:")
    print(sorted(hw_assign_fields + ["CABINET", "STATION", "SLOT", "CHANNEL", "RT", "TYPE", "MLFB", "change_token", "KKS", "SIGDEF"]))
    print("")
    print("IOs for following modules will be checked:")
    print(sorted(config_dict['modules']))
    print("")

    wb_obj = load_workbook(file_hw_assign)
    sheets = wb_obj.sheetnames

    for sheet in sheets:
        if sheet.startswith("HW assignment"): break

    sheet_obj = wb_obj[sheet]

    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row
    
    def extact_col_vals(max_row, col, lst): #this function has no return value. It only has side effect.
        for i in range(1, max_row + 1):
            _val = sheet_obj.cell(row = i, column = col).value
            if _val == None: lst.append(_val)
            else: lst.append(str(_val).strip())

    cabinet_col, station_col, slot_col, channel_col, rt_col, type_col, mlfb_col, change_token_col = None, None, None, None, None, None, None, None
    kks_col, sigdef_col = None, None

    varName_col = [] # list of dynamically generated variable names
    varName_col_list = [] # list of dynamically generated variable names
    for idx, field in enumerate(hw_assign_fields):
        varName = "_var_field" + str(idx) + "_col"
        varNameList = varName + "_list"
        locals()[varName] = None
        locals()[varNameList] = []
        varName_col.append(varName)
        varName_col_list.append(varNameList)
        for i in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row = 1, column = i)
            if cell_obj.value == field: locals()[varName] = i
         
    # pprint(locals()[varName_col[0]])
    # pprint(locals()[varName_col[1]])
    # pprint(locals()[varName_col[2]])
    # pprint(locals()[varName_col[3]])
    # pprint(locals()[varName_col[4]])
    # pprint(locals()[varName_col[5]])


    for i in range(1, max_col + 1):
        cell_obj = sheet_obj.cell(row = 1, column = i)
        if cell_obj.value == "CABINET": cabinet_col = i
        elif cell_obj.value == "STATION": station_col = i
        elif cell_obj.value == "SLOT": slot_col = i
        elif cell_obj.value == "CHANNEL": channel_col = i
        elif cell_obj.value == "RT": rt_col = i
        elif cell_obj.value == "TYPE": type_col = i
        elif cell_obj.value == "MLFB": mlfb_col = i
        elif cell_obj.value == "change_token": change_token_col = i
        elif cell_obj.value == "KKS": kks_col = i
        elif cell_obj.value == "SIGDEF": sigdef_col = i
            


    assert cabinet_col != None, 'column "CABINET" missing in HW_assignment_list, program aborted.'
    assert station_col != None, 'column "STATION" missing in HW_assignment_list, program aborted.'
    assert slot_col != None, 'column "SLOT" missing in HW_assignment_list, program aborted.'
    assert channel_col != None, 'column "CHANNEL" missing in HW_assignment_list, program aborted.'
    assert rt_col != None, 'column "RT" missing in HW_assignment_list, program aborted.'
    assert type_col != None, 'column "TYPE" missing in HW_assignment_list, program aborted.'
    assert mlfb_col != None, 'column "MLFB" missing in HW_assignment_list, program aborted.'
    assert change_token_col != None, 'column "change_token" missing in HW_assignment_list, program aborted.'
    assert kks_col != None, 'column "KKS" missing in HW_assignment_list, program aborted.'
    assert sigdef_col != None, 'column "SIGDEF" missing in HW_assignment_list, program aborted.'

    for i in range(len(hw_assign_fields)):
        assert locals()[varName_col[i]] != None, f'column "{hw_assign_fields[i]}" missing in HW_assignment_list, program aborted.'
        
        
    cabinet_col_list, station_col_list, slot_col_list, channel_col_list, rt_col_list, type_col_list, mlfb_col_list, change_token_col_list = [], [], [], [], [], [], [], []
    kks_col_list, sigdef_col_list = [], []

    extact_col_vals(max_row, cabinet_col, cabinet_col_list)
    extact_col_vals(max_row, station_col, station_col_list)
    extact_col_vals(max_row, slot_col, slot_col_list)
    extact_col_vals(max_row, channel_col, channel_col_list)
    extact_col_vals(max_row, rt_col, rt_col_list)
    extact_col_vals(max_row, type_col, type_col_list)
    extact_col_vals(max_row, mlfb_col, mlfb_col_list)
    extact_col_vals(max_row, change_token_col, change_token_col_list)
    extact_col_vals(max_row, kks_col, kks_col_list)
    extact_col_vals(max_row, sigdef_col, sigdef_col_list)

    for i in range(len(hw_assign_fields)):
        extact_col_vals(max_row, locals()[varName_col[i]], locals()[varName_col_list[i]])
        
    # pprint(locals()[varName_col_list[0]])
    # pprint(locals()[varName_col_list[1]])
    # pprint(locals()[varName_col_list[2]])
    # pprint(locals()[varName_col_list[3]])
    # pprint(locals()[varName_col_list[4]])
    # pprint(locals()[varName_col_list[5]])
        
    io_table_dict = {}
    tot_processed_rows = 0
    for i in range(1, len(mlfb_col_list)):

        mlfb = mlfb_col_list[i]
        kks = kks_col_list[i]
        chg_token = change_token_col_list[i]
        cabinet = cabinet_col_list[i]
        station = station_col_list[i]
        slot = slot_col_list[i]
        channel = channel_col_list[i]
        rt = rt_col_list[i]
        sigdef = sigdef_col_list[i]
        
        if mlfb == None: continue
        if not (mlfb in config_dict['modules'].keys()): continue
        
        if '$' in str(chg_token) or 'spare' in str(chg_token).lower(): continue
            
        if kks == None: continue
        if kks == "": continue
        
        tot_processed_rows = tot_processed_rows + 1
            
        if not re.match("^[0-9]", kks): kks = re.sub("([^a-zA-Z]*)(.*)", prefix + r"\2", kks, 1)
            
        # append "HW" to input signals
        assert sigdef != None and sigdef != "", f'Row-{i+1} of HW_assignment_list has undefined "SIGDEF" field.'
        if rt == "AI" or rt == "DI": sigdef = sigdef + "HW"
            
        assert (cabinet in config_cabinet), f'Row-{i+1} of HW_assignment_list has undefined "CABINET" field, program aborted. Expected values are {config_cabinet}.'
        if not re.match("^[0-9]", cabinet): cabinet = re.sub("([^a-zA-Z]*)(.*)", prefix + r"\2", cabinet, 1)
        
        assert (station in config_station), f'Row-{i+1} of HW_assignment_list has unexpected "STATION" field, program aborted. Expected values are {config_station}.'
        
        try: slot = str(int(slot))
        except: raise ValueError (f'Row-{i+1} of HW_assignment_list has unexpected "SLOT" field, program aborted. Expected values are {config_slot}.')
        assert (slot in config_slot), f'Row-{i+1} of HW_assignment_list has unexpected "SLOT" field, program aborted. Expected values are {config_slot}.'
        int_slot = int(slot)

        try: channel = str(int(channel))
        except: raise ValueError (f'Row-{i+1} of HW_assignment_list has unexpected "CHANNEL" field, program aborted. Expected values are {config_channel}.')
        assert (channel in config_channel), f'Row-{i+1} of HW_assignment_list has unexpected "CHANNEL" field, program aborted. Expected values are {config_channel}.'
        int_channel = int(channel)
        
        assert (rt in config_rt), f'Row-{i+1} of HW_assignment_list has unexpected "RT" field, program aborted. Expected values are {config_rt}.'

        attribute_addr_base = tagname_t3k_hw_proxy(cabinet, station, slot)

        attribute_addr_base = attribute_addr_base + "::" + mlfb
        
        for k,v in config_dict['modules'][mlfb][rt].items():
            
            attribute_addr = attribute_addr_base + "::" + k.replace("xx_", f"{int_channel:02}_")
            assert (attribute_addr not in io_table_dict), f'IO address at Row-{i+1} of HW_assignment_list is duplicate or this can be a issue with toml config file, program aborted.'
            
            if isinstance(v, dict): # is instance of dict means: it needs to make judgement on what field to lookup in HW assignment list
                lookup_v = v['lookup'].replace("@", "").strip()
                
                # check ["CABINET", "STATION", "SLOT", "CHANNEL", "RT", "TYPE", "MLFB", "change_token", "KKS", "SIGDEF"]
                if lookup_v == "CABINET": lookup_v = cabinet
                elif lookup_v == "STATION": lookup_v = station
                elif lookup_v == "SLOT": lookup_v = slot
                elif lookup_v == "CHANNEL": lookup_v = channel
                elif lookup_v == "RT": lookup_v = rt
                elif lookup_v == "TYPE": lookup_v = type_col_list[i]
                elif lookup_v == "MLFB": lookup_v = mlfb
                elif lookup_v == "change_token": lookup_v = chg_token
                elif lookup_v == "KKS": lookup_v = kks
                elif lookup_v == "SIGDEF": lookup_v = sigdef
                
                else:
                    idx_v = hw_assign_fields.index(lookup_v)
                    lookup_v = locals()[varName_col_list[idx_v]][i]
                
                # if lookuped value is not defined in the dict, then, use default value
                if not lookup_v in v.keys(): v = v["default"]
                else: v = v[lookup_v]      
                
            elif v.startswith('@'): # starts with '@' means: it needs to be looked up in HW assignment list
                v = v.replace('@', "").strip()
                
                # check ["CABINET", "STATION", "SLOT", "CHANNEL", "RT", "TYPE", "MLFB", "change_token", "KKS", "SIGDEF"]
                if v == "CABINET": v = cabinet
                elif v == "STATION": v = station
                elif v == "SLOT": v = slot
                elif v == "CHANNEL": v = channel
                elif v == "RT": v = rt
                elif v == "TYPE": v = type_col_list[i]
                elif v == "MLFB": v = mlfb
                elif v == "change_token": v = chg_token
                elif v == "KKS": v = kks
                elif v == "SIGDEF": v = sigdef
                
                else:
                    idx_v = hw_assign_fields.index(v)
                    v = locals()[varName_col_list[idx_v]][i]
                    if v == None or v.strip() == "": v = "check?"
                    
            else: pass  # value of v is used as it is from the toml config file
            
            assert isinstance(v, str), "Value not a string. Check toml config file"
            
            io_table_dict[attribute_addr] = v
            
    print("Total number of rows processed in HW_assignment_list:")
    print(tot_processed_rows)
    print("")
    
    if file_hw_assign_out != None:
        # writing to a new excel file
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.cell(row=1, column=1).value = "attribute_addr"
        worksheet.cell(row=1, column=2).value = "attribute_val"
        for row, (key, value) in enumerate(sorted(io_table_dict.items()), start=2):
            worksheet.cell(row=row, column=1).value = key
            worksheet.cell(row=row, column=2).value = value
        workbook.save(file_hw_assign_out)
        
    return io_table_dict
    

def io_data_t3000(file_t3000):
    module_list = config_dict['modules']
    
    wb_obj = load_workbook(file_t3000)
    sheets = wb_obj.sheetnames

    for sheet in sheets:
        if sheet.startswith("Data"): break

    sheet_obj = wb_obj[sheet]

    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row
    
    def extact_col_vals(col, lst): #this function has no return value. It only has side effect.
        nonlocal max_row
        for i in range(1, max_row + 1):
            _val = sheet_obj.cell(row = i, column = col).value
            if _val == None: lst.append(_val)
            else: lst.append(str(_val).strip())

    # def extact_col_vals(col, lst): # this function has no return value. It only has side effect.
    #     with open(file_t3000, newline='', encoding='utf-16') as csvfile:
    #         reader = csv.reader(csvfile, delimiter=';')
    #         transposed = list(zip(*reader)) # Transpose the CSV file using the zip() function
    #         lst.extend(transposed[col]) # Get the column with index col as a list
    
    
    t3000_port_fields = re.findall('\<(.*?)\>', config_dict_str)
    t3000_port_fields = list(set(t3000_port_fields) - set(['Tag- Name', 'Symbol- Type', 'Port- Type', 'Signal', 'Value']))
    varName_col = [] # list of dynamically generated variable names
    varName_col_list = [] # list of dynamically generated variable names
    varName_col_list_temp = []
    for idx, field in enumerate(t3000_port_fields):
        varName = "_var_field" + str(idx) + "_col"
        varNameList = varName + "_list"
        varNameList_temp = varNameList + "_temp"
        locals()[varName] = None
        locals()[varNameList] = []
        locals()[varNameList_temp] = []
        varName_col.append(varName)
        varName_col_list.append(varNameList)
        varName_col_list_temp.append(varNameList_temp)

        # with open(file_t3000, newline='', encoding='utf-16') as csvfile:
        #     reader = csv.reader(csvfile, delimiter=';')
        #     header = next(reader) # Get the first row of the CSV file
        #     # Find the index of the column with the header "Column Name"
        #     try: locals()[varName] = header.index(field)
        #     except ValueError: pass

        for i in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row = 1, column = i)
            if cell_obj.value == field: locals()[varName] = i
                
    for i in range(len(t3000_port_fields)):
        assert locals()[varName_col[i]] != None, f'column "{t3000_port_fields[i]}" missing in T3000 data, program aborted.'
    
    tagName_col, symbolType_col, portName_col, signal_col, value_col = None, None, None, None, None

    # with open(file_t3000, newline='', encoding='utf-16') as csvfile:
    #     reader = csv.reader(csvfile, delimiter=';')
    #     header = next(reader) # Get the first row of the CSV file
    #     # Find the index of the column with the header "Column Name"
    #     try: tagName_col = header.index("Tag- Name")
    #     except ValueError: pass
    #     try: symbolType_col = header.index("Symbol- Type")
    #     except ValueError: pass
    #     try: portName_col = header.index("Port- Name")
    #     except ValueError: pass
    #     try: signal_col = header.index("Signal")
    #     except ValueError: pass
    #     try: value_col = header.index("Value")
    #     except ValueError: pass

    for i in range(1, max_col + 1):
        cell_obj = sheet_obj.cell(row = 1, column = i)
        if cell_obj.value == "Tag- Name": tagName_col = i
        elif cell_obj.value == "Symbol- Type": symbolType_col = i
        elif cell_obj.value == "Port- Name": portName_col = i
        elif cell_obj.value == "Signal": signal_col = i
        elif cell_obj.value == "Value": value_col = i
            
    assert tagName_col != None, 'column "Tag- Name" missing in T3000 data, program aborted.'
    assert symbolType_col != None, 'column "Symbol- Type" missing in T3000 data, program aborted.'
    assert portName_col != None, 'column "Port- Name" missing in T3000 data, program aborted.'
    assert signal_col != None, 'column "Signal" missing in T3000 data, program aborted.'
    assert value_col != None, 'column "Value" missing in T3000 data, program aborted.'
    
    
    for i in range(len(t3000_port_fields)):
        extact_col_vals(locals()[varName_col[i]], locals()[varName_col_list[i]])
        
    tagName_col_list, symbolType_col_list, portName_col_list, signal_col_list, value_col_list  = [], [], [], [], []
    tagName_col_list_temp, symbolType_col_list_temp, portName_col_list_temp, signal_col_list_temp, value_col_list_temp = [], [], [], [], []
    
    extact_col_vals(tagName_col, tagName_col_list)
    extact_col_vals(symbolType_col, symbolType_col_list)
    extact_col_vals(portName_col, portName_col_list)
    extact_col_vals(signal_col, signal_col_list)
    extact_col_vals(value_col, value_col_list)
    
    for i in range(len(tagName_col_list)):
        if i != 0: # index 0 is title for the list and must be retained
            # check if module defined in config file and portName format is "xx_"
            if (symbolType_col_list[i] not in module_list) or (re.search("[0-9]{2}_", portName_col_list[i]) == None): continue
        tagName_col_list_temp.append(tagName_col_list[i])
        symbolType_col_list_temp.append(symbolType_col_list[i])
        portName_col_list_temp.append(portName_col_list[i])
        signal_col_list_temp.append(signal_col_list[i])
        value_col_list_temp.append(value_col_list[i])
        
        for k in range(len(varName_col_list)): locals()[varName_col_list_temp[k]].append(locals()[varName_col_list[k]][i])
            
    tagName_col_list, symbolType_col_list, portName_col_list, signal_col_list, value_col_list = tagName_col_list_temp, symbolType_col_list_temp, portName_col_list_temp, signal_col_list_temp, value_col_list_temp

    tagName_col_list_temp, symbolType_col_list_temp, portName_col_list_temp, signal_col_list_temp, value_col_list_temp = [], [], [], [], []
    
    for k in range(len(varName_col_list)):
        locals()[varName_col_list[k]] = locals()[varName_col_list_temp[k]]
        locals()[varName_col_list_temp[k]] = []
   
    # pprint(signal_col_list)
    used_ch_dict = {}
    for i in range(1, len(tagName_col_list)): # skipping index 0 as it is the title of the list
        if signal_col_list[i].lower() == 'false': continue
        portName = portName_col_list[i]
        match = re.search('^([A-Z]{2})([0-9]{2})_PV$', portName)
        if match == None: continue
        tageName = tagName_col_list[i]
        symbolType = symbolType_col_list[i]
        if tageName not in used_ch_dict.keys(): used_ch_dict.update({tageName: {'symbolType': symbolType}})
        rt = match.group(1)
        ch = match.group(2)
        if rt not in used_ch_dict[tageName].keys(): used_ch_dict[tageName].update({rt : []})
        used_ch_dict[tageName][rt].append(ch)
    # pprint(used_ch_dict)
    
    used_port_dict = {}
    for key, val in used_ch_dict.items():
        symbolType = val['symbolType']
        if key not in used_port_dict.keys(): used_port_dict.update({key: {}}) # key = '11CPA02FB|SLOT 30'
        for rt, channels in val.items():
            if rt == 'symbolType': continue
            portAttrs_xx = config_dict['modules'][symbolType][rt].keys() # [AIxx_PV<>,...]
            for channel in channels:
                portAttrs = [x.replace("xx_", channel + "_") for x in portAttrs_xx]
                for portAttr in portAttrs:
                    x = portAttr.split("<")
                    portName = x[0]
                    attr = x[1].replace(">", "")
                    if portName not in used_port_dict[key].keys(): used_port_dict[key].update({portName: []})
                    used_port_dict[key][portName].append(attr)            
    # pprint(used_port_dict)
    
    for i in range(len(tagName_col_list)):
        tagName = tagName_col_list[i]
        if i != 0: # index 0 is title for the list and must be retained
            if tagName not in used_port_dict.keys(): continue
        portName = portName_col_list[i]
        if i != 0: # index 0 is title for the list and must be retained
            if portName not in used_port_dict[tagName].keys(): continue
            
        tagName_col_list_temp.append(tagName)
        symbolType_col_list_temp.append(symbolType_col_list[i])
        portName_col_list_temp.append(portName)
        signal_col_list_temp.append(signal_col_list[i])
        value_col_list_temp.append(value_col_list[i])
        
        for k in range(len(varName_col_list)): locals()[varName_col_list_temp[k]].append(locals()[varName_col_list[k]][i])
            
    tagName_col_list, symbolType_col_list, portName_col_list, signal_col_list, value_col_list = tagName_col_list_temp, symbolType_col_list_temp, portName_col_list_temp, signal_col_list_temp, value_col_list_temp

    tagName_col_list_temp, symbolType_col_list_temp, portName_col_list_temp, signal_col_list_temp, value_col_list_temp = [], [], [], [], []
    
    # delete ", Ch x disabled" from "Value" column
    value_col_list = [re.sub(", Ch [0-9] disabled", "", x) for x in value_col_list]
    
    for k in range(len(varName_col_list)):
        locals()[varName_col_list[k]] = locals()[varName_col_list_temp[k]]
        locals()[varName_col_list_temp[k]] = []
    
    t3000_dict = {}
    for i in range(1, len(tagName_col_list)): # skipping index 0 as it is the title of the list
        tagName = tagName_col_list[i]
        symbolType = symbolType_col_list[i]
        portName = portName_col_list[i]
        
        for attr in used_port_dict[tagName][portName]:
            attribute_addr = f"{tagName}::{symbolType}::{portName}<{attr}>"
            if attr in 'Tag- Name':
                t3000_dict[attribute_addr] = tagName_col_list[i]
                continue
            if attr == 'Symbol- Type':
                t3000_dict[attribute_addr] = symbolType_col_list[i]
                continue
            if attr == 'Port- Type':
                t3000_dict[attribute_addr] = portName_col_list[i]
                continue
            if attr == 'Signal':
                t3000_dict[attribute_addr] = signal_col_list[i]
                continue
            if attr == 'Value':
                t3000_dict[attribute_addr] = value_col_list[i]
                continue  
            varName_col_list_len = len(varName_col_list)
            for k in range(varName_col_list_len):
                if locals()[varName_col_list[k]][0] == attr: break
                assert k != varName_col_list_len - 1, "attribute list not found"
            t3000_dict[attribute_addr] = locals()[varName_col_list[k]][i]
            
    if file_t3000_assign_out != None:
        # writing to a new excel file
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.cell(row=1, column=1).value = "attribute_addr"
        worksheet.cell(row=1, column=2).value = "attribute_val"
        for row, (key, value) in enumerate(sorted(t3000_dict.items()), start=2):
            worksheet.cell(row=row, column=1).value = key
            worksheet.cell(row=row, column=2).value = value
        workbook.save(file_t3000_assign_out)
    
    return t3000_dict
    

def diff_check(t3000_dict, io_table_dict):
    print("Generating difference.")
    diff = dict(deepdiff.DeepDiff(t3000_dict, io_table_dict, verbose_level=2))
    for k, v in diff.items():
        if type(v) == deepdiff.model.PrettyOrderedSet:
            diff[k] = list(v)

    def is_eq_float(x1, x2):
        try: return float(x1) == float(x2)
        except ValueError: return False

    dict_added = diff.get('dictionary_item_added', {}) # added in io_table
    dict_removed = diff.get('dictionary_item_removed', {}) # removed from io_table
    dict_changed = diff.get('values_changed', {})
    if dict_changed != {}:
        dict_changed = {k : v for (k, v) in dict_changed.items() if not is_eq_float(v["old_value"], v["new_value"])}

    result_header = ['Tag- Name', 'Symbol- Type', 'Port- Name <Attribute>', '', 'Change remark', 'T3000 (old)', 'HW_assignment_list (new)']
    list_result_diff = []

    for key, value in dict_added.items():
        sub_list = re.findall('(?<=\[\').*?(?=\'\])', key)
        attribute_addr_list = sub_list[0].split('::')
        tagName = attribute_addr_list[0]
        symbolType = attribute_addr_list[1]
        portNameWithAttr = attribute_addr_list[2]
        chgRemark = 'new'
        oldVal = ""
        newVal = value
        list_result_diff.append([tagName, symbolType, portNameWithAttr, '→', chgRemark, oldVal, newVal])

    for key, value in dict_removed.items():
        sub_list = re.findall('(?<=\[\').*?(?=\'\])', key)
        attribute_addr_list = sub_list[0].split('::')
        tagName = attribute_addr_list[0]
        symbolType = attribute_addr_list[1]
        portNameWithAttr = attribute_addr_list[2]
        chgRemark = '$'
        oldVal = value
        newVal = ""
        list_result_diff.append([tagName, symbolType, portNameWithAttr, '→', chgRemark, oldVal, newVal])

    for key, value in dict_changed.items():
        sub_list = re.findall('(?<=\[\').*?(?=\'\])', key)
        attribute_addr_list = sub_list[0].split('::')
        tagName = attribute_addr_list[0]
        symbolType = attribute_addr_list[1]
        portNameWithAttr = attribute_addr_list[2]
        chgRemark = 'mod'
        oldVal = value['old_value']
        newVal = value['new_value']
        list_result_diff.append([tagName, symbolType, portNameWithAttr, '→', chgRemark, oldVal, newVal])
    
    return [result_header] + sorted(list_result_diff, key = lambda x: (x[0], x[2]))
    

def write_diff():
    if file_t3000_2 == None:
        print("Processing T3000 data file1.\n")
        t3000_dict = io_data_t3000(file_t3000_1)
    else:
        print("Processing T3000 data file1 and file2.\n")
        t3000_dict = io_data_t3000(file_t3000_1) | io_data_t3000(file_t3000_2)
        
    print("Processing HW_assignment_list.\n")
    io_table_dict = io_data_hw_assignment()
    
    list_result_diff = diff_check(t3000_dict, io_table_dict)

    print("Writing results.")
    wb = openpyxl_wb() # create workbook object.
    ws = wb.active # create worksheet object.
    for row in list_result_diff:
        ws.append(row) # adds values to cells, each list is a new row.
    wb.save(file_diff) # save to excel file.

